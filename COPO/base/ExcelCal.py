from django.http import HttpResponse
from django.shortcuts import render
import openpyxl
from io import BytesIO
import formulas
from openpyxl import load_workbook
import pandas as pd
from tempfile import NamedTemporaryFile
from tempfile import TemporaryDirectory
import os
import win32com.client as w3c

class ExcelCal:
   
    
    #------------------------
    def sumAndPer_cal(self,start,end,store,persent,outOff):
        #global sheet
        #global h
        sheet = self.sheet
        h = self.h
        l = len(h)
        for i, cellobj in enumerate(sheet.iter_rows(min_row=self.row_offset), self.row_offset):
            sheet["${}{}".format(store,i)] = "=SUM(${0}{2}:{1}{2})".format(start,end,i)
            sheet["${}{}".format(persent,i)] = "=${}{}/{}*100".format(store,i,outOff)

    def per_cal(self,seed,store,outOff):
        #global sheet
        #global h
        sheet = self.sheet
        h = self.h
        for i, celobj in enumerate(sheet.iter_rows(min_row=self.row_offset),self.row_offset):
            sheet["${}{}".format(store,i)] = "=${}{}/{}*100".format(seed,i,outOff)
    
    def COtable(self,row,col,start,CO=50):
        #global sheet
        #global h
        sheet = self.sheet
        h = self.h
        l = len(sheet['A'])
        #print(l)
        out = ["CO's","Criteria","Count","%"]
        for i in range(0,4):
            sheet.cell(row=row,column=col+i).value = out[i]
        for i in range(0,len(h)):
            C = ord(start)+i*2
            S = ord('A')
            #=COUNTIF(O9:O77,">=50")
            out = [h[i],"No. of students scored >={}".format(CO),
                '''=COUNTIF('''+chr(C)+"{}:".format(self.row_offset)+chr(C)+'''{},">={}")'''.format(l,CO),
                "="+chr(S+col+1)+"{}/{}*100".format(row+i+1,l-self.row_offset+1)]
            for j in range(0,4):
                sheet.cell(row=row+i+1,column=col+j).value = out[j]

    def COtable1(self,row,col,start):
        #global sheet
        #global h
        sheet = self.sheet
        h = self.h
        l = len(sheet['A'])
        #print(l)
        out = ["CO's","Criteria","Count","%"]
        for i in range(0,4):
            sheet.cell(row=row,column=col+i).value = out[i]
        for i in range(0,len(h)):
            C = self.getC(start,i)
            S = ord('A')
            #=COUNTIF(O9:O77,">=50")
            out = [h[i],"No. of students scored >=60",
                '''=COUNTIF('''+chr(C)+"{}:".format(self.row_offset)+chr(C)+'''{},">=60")'''.format(l),
                "="+chr(S+col+1)+"{}/{}*100".format(row+i+1,l-self.row_offset+1)]
            for j in range(0,4):
                sheet.cell(row=row+i+1,column=col+j).value = out[j]
    def getC(self,s,i):
        #global x,sheet
        x = self.x
        sheet = self.sheet
        for j in range(1,sheet.max_column+1):
            if(sheet.cell(self.row_offset-2,j).value == '%'):
                if i == 0:
                    return ord(s)+j-1
                else:
                    i = i-1

    def attainmentLevel1(self,row,col,seed):
        #global wb
        #global sheet
        #global h
        wb = self.wb
        sheet = self.wb
        h = self.h
        l = len(sheet['A'])
        start = l - len(h)
        #print(h)
        out = ["Level","Attainment Level"]

        out.extend(h)
        #print(out)
        for i in range(0,len(h)+2):
            sheet.cell(row=row,column=col+i).value = out[i]
        per = ["55 % of studens scored >=50","75 % of studens scored >=50","80 % of studens scored >=50"]
        per1 = [55,75,80,100]
        for i in range(1,4):
            out =[i,
                per[i-1],
                "=if("+seed+"{0}>={1},if(".format(start+1,per1[i-1])+seed+"{0}<{1},{2},0),0)".format(start+1,per1[i],i),
                "=if("+seed+"{0}>={1},if(".format(start+2,per1[i-1])+seed+"{0}<{1},{2},0),0)".format(start+2,per1[i],i),
                "=if("+seed+"{0}>={1},if(".format(start+3,per1[i-1])+seed+"{0}<{1},{2},0),0)".format(start+3,per1[i],i),
                "=if("+seed+"{0}>={1},if(".format(start+4,per1[i-1])+seed+"{0}<{1},{2},0),0)".format(start+4,per1[i],i),
                "=if("+seed+"{0}>={1},if(".format(start+5,per1[i-1])+seed+"{0}<{1},{2},0),0)".format(start+5,per1[i],i),
                "=if("+seed+"{0}>={1},if(".format(start+6,per1[i-1])+seed+"{0}<{1},{2},0),0)".format(start+6,per1[i],i),]

            for j in range(0,len(h)+2):
                sheet.cell(row=row+i,column=col+j).value = out[j]
        
        w=[50,50]
        if "Final_attainment" not in wb.sheetnames:
            wb.create_sheet("Final_attainment")
        sh = wb["Final_attainment"]    
        st = ord('A')
        l = len(sh['A'])
        sh.cell(l+1,1).value = sheet.title
        for i in range(0,len(h)):
            #print(sheet.cell(row=row,column=col+2+i).value)
            c = chr(st+col+1+i)
            if h[i] != "ALL":
                col1 = int((h[i])[-1])
                #print(col1)
                #print(l)
                sh.cell(1,col1+2).value=h[i]
                sh.cell(l+1,col1+2).value = "=SUM('"+str(sheet.title)+"'!"+c+"{}".format(row+1)+":"+c+"{}".format(row+3)+")"
            else:
                #print(sh.max_column)
                for j in range(3,sh.max_column+1):
                    sh.cell(l+1,j).value = "=SUM('"+str(sheet.title)+"'!"+c+"{}".format(row+1)+":"+c+"{}".format(row+3)+")"
                sh.cell(1,1).value="Asssessment tools used" 
                sh.cell(1,2).value="Weightage"    
                for i in range(0,2):    
                    sh.cell(i+2,2).value="{}%".format(w[i])
                sh.cell(4,2).value="total"
                st = ord('A') - 1
                for i in range(3,9):
                    formula = "="
                    for j in range(2,4):
                        formula = formula + chr(st+i)+"{}*B{}+".format(j,j)
                    formula = formula[0:-1]
                    sh.cell(4,i).value=formula
                sh.cell(3,9).value="AVERAGE"    
                sh.cell(4,9).value="=AVERAGE(C4:H4)"


    def attainmentLevel(self,row,col,seed):
        #global wb
        #global sheet
        #global h
        #global temp
        global per1,w
        
        wb = self.wb
        sheet = self.sheet
        h = self.h
        self.temp = h.copy()
        l = len(sheet['A'])
        start = l - len(h)
        #print(h)
        out = ["Level","Attainment Level"]

        out.extend(h)
        #print(out)
        for i in range(0,len(h)+2):
            sheet.cell(row=row,column=col+i).value = out[i]
        #per1 = [40,60,80,100]
        per = [f"{per1[0]} % of studens scored >=50",f"{per1[1]} % of studens scored >=50",f"{per1[2]} % of studens scored >=50"]
        
        for i in range(1,4):
            out =[i,
                per[i-1],
                "=if("+seed+"{0}>={1},if(".format(start+1,per1[i-1])+seed+"{0}<{1},{2},0),0)".format(start+1,per1[i],i),
                "=if("+seed+"{0}>={1},if(".format(start+2,per1[i-1])+seed+"{0}<{1},{2},0),0)".format(start+2,per1[i],i),
                "=if("+seed+"{0}>={1},if(".format(start+3,per1[i-1])+seed+"{0}<{1},{2},0),0)".format(start+3,per1[i],i),]

            for j in range(0,len(h)+2):
                sheet.cell(row=row+i,column=col+j).value = out[j]
        
        #w=[20,20,10,50]
        if "Final_attainment" not in wb.sheetnames:
            wb.create_sheet("Final_attainment")
        sh = wb["Final_attainment"]    
        st = ord('A')
        l = len(sh['A'])
        sh.cell(l+1,1).value = sheet.title
        for i in range(0,len(h)):
            #print(sheet.cell(row=row,column=col+2+i).value)
            c = chr(st+col+1+i)
            if h[i] != "ALL":
                col1 = int((h[i])[-1])
                #print(col1)
                #print(l)
                sh.cell(1,col1+2).value=h[i]
                sh.cell(l+1,col1+2).value = "=SUM('"+str(sheet.title)+"'!"+c+"{}".format(row+1)+":"+c+"{}".format(row+3)+")"
            else:
                #print(sh.max_column)
                for j in range(3,sh.max_column+1):
                    sh.cell(l+1,j).value = "=SUM('"+str(sheet.title)+"'!"+c+"{}".format(row+1)+":"+c+"{}".format(row+3)+")"
                sh.cell(1,1).value="Asssessment tools used" 
                sh.cell(1,2).value="Weightage"    
                for i in range(0,4):    
                    sh.cell(i+2,2).value="{}%".format(w[i])
                sh.cell(6,2).value="total"
                st = ord('A') - 1
                for i in range(3,9):
                    formula = "="
                    for j in range(2,6):
                        formula = formula + chr(st+i)+"{}*B{}+".format(j,j)
                    formula = formula[0:-1]
                    sh.cell(6,i).value=formula
                sh.cell(5,9).value="AVERAGE"    
                sh.cell(6,9).value="=AVERAGE(C6:H6)"
        
        

    #print(wb.sheetnames)

    #print(len(sheet['A']))

    #--------IA----------
    #print(sheet["A"])
    def IA(self):
        #global wb,sheet,h,row_offset
        self.h = list()
        for row in self.sheet.iter_rows(min_row=self.row_offset-1,
                                max_row=self.row_offset-1):
            for cell in row:
                if(cell.value!=None):
                    if(cell.value not in self.h):
                        self.h.append(cell.value)
        #print(h)

        if(len(self.h)==2):
            self.sheet["N{}".format(self.row_offset-1)] = self.h[0]
            self.sheet["O{}".format(self.row_offset-1)] = self.h[0]+" %"
            self.sumAndPer_cal('C','H','N','O',10)
            self.sheet["P{}".format(self.row_offset-1)] = self.h[1]
            self.sheet["Q{}".format(self.row_offset-1)] = self.h[1]+" %"
            self.sumAndPer_cal('I','L','P','Q',10)
        elif(len(self.h)==3):
            self.sheet["N{}".format(self.row_offset-1)] = self.h[0]
            self.sheet["O{}".format(self.row_offset-1)] = self.h[0]+" %"
            self.sumAndPer_cal('C','H','N','O',10)
            self.sheet["P{}".format(self.row_offset-1)] = self.h[1]
            self.sheet["Q{}".format(self.row_offset-1)] = self.h[1]+" %"
            self.sumAndPer_cal('I','J','P','Q',5)
            self.sheet["R{}".format(self.row_offset-1)] = self.h[2]
            self.sheet["S{}".format(self.row_offset-1)] = self.h[2]+" %"
            self.sumAndPer_cal('K','L','R','S',5)
        self.COtable(len(self.sheet['A'])+2,2,'O')
        self.attainmentLevel(len(self.sheet['B'])+2,2,'E')

    #--------IA------------
    def Assignment(self):
        #global wb,sheet,h,row_offset
        self.row_offset=7
        temp = self.h[0]
        temp = temp[0:-1]+'6'
        self.h =[temp]
        self.sheet["F{}".format(self.row_offset-1)] = "%"
        self.per_cal("E","F",5)
        self.COtable(len(self.sheet['A'])+2,2,'F')
        self.attainmentLevel(len(self.sheet['B'])+2,2,'E')

    def ESE(self):
        #global wb,sheet,h,row_offset
        self.row_offset=8
        self.h =["ALL"]
        self.sheet["C{}".format(self.row_offset-1)] = "%"
        self.per_cal("B","C",80)
        self.COtable(len(self.sheet['A'])+2,2,'C')
        self.attainmentLevel(len(self.sheet['B'])+2,2,'E')

    #---------------------------------------------------------------------------------------------------------------------------

    def ExpAttainment(self):
        global wb,sheet,h,row_offset
        self.row_offset=9
        self.h = list()
        for row in sheet.iter_rows(min_row=self.row_offset-1,
                                max_row=self.row_offset-1):
            for cell in row:
                if(cell.value!=None):
                    if(cell.value not in h):
                        h.append(cell.value)
        self.h=h[1:]
        self.COtable1(len(sheet['A'])+2,2,'A')
        self.attainmentLevel1(len(sheet['B'])+2,2,'E')
    def LabUniversity(self):
        global wb,sheet,h,row_offset
        self.row_offset=8
        self.h =["ALL"]
        self.sheet["F{}".format(self.row_offset-1)] = "%"
        self.per_cal("E","F",25)
        self.COtable(len(sheet['A'])+2,2,'F',60)
        self.attainmentLevel1(len(sheet['B'])+2,2,'E')

    #------------------------

def cal(in_file,CourseOutcomeT,AttainmentT,type1):
    global wb,sheet,h,temp,per1,w
    per1 = CourseOutcomeT
    w = AttainmentT
    excel_file = in_file

    # you may put validations here to check extension or file size
    
    #wb = load_workbook(filename="Test1.xlsx")
    Ec = ExcelCal()
    Ec.row_offset = 9
    Ec.wb = openpyxl.load_workbook(excel_file)
    #wb = openpyxl.load_workbook(excel_file)
    
    #--------------------------------------------------------
    #--------------------------------------------------------
    #--------------------------------------------------------
    
    x = type1
    if(x=="Theory"):
        Ec.sheet = Ec.wb['IA-1']
        Ec.IA()
        Ec.sheet = Ec.wb['IA-2']
        Ec.IA()
        Ec.sheet = Ec.wb['Assignment']
        Ec.Assignment()
        Ec.sheet = Ec.wb['ESE']
        Ec.ESE()
        Ec.h=Ec.temp.copy()
    else:
        Ec.sheet = Ec.wb['Exp Attainment']
        Ec.ExpAttainment()
        Ec.sheet = Ec.wb['Lab University']
        Ec.LabUniversity()

    wb = Ec.wb
    import pythoncom
    import struct
    import json
    pythoncom.CoInitialize()
    name='A.xlsx'
    wb.save(name)
    wb.close()
    ex_file = w3c.DispatchEx('Excel.Application')
    wb = ex_file.Workbooks.Open(os.path.abspath(name))
    wb.RefreshAll()
    wb.Save()
    ex_file.Quit()
    B = BytesIO()
    with open(os.path.abspath(name),'rb') as fh:
        B=BytesIO(fh.read())
    #wb = load_workbook(name,data_only=True)
    #wb.save(B)
    int_list = struct.unpack(f"{len(B.getvalue())}B",B.getvalue())
    json_string = json.dumps(int_list)
    

    #return response
    return json_string , name


